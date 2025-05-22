import os
import psycopg2
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta
import sys
import copy

def main():
    # 1. Obtener credenciales y ruta del archivo
    db_name = os.environ.get('DB_NAME')
    db_user = os.environ.get('DB_USER')
    db_password = os.environ.get('DB_PASSWORD')
    db_host = os.environ.get('DB_HOST')
    db_port = os.environ.get('DB_PORT')
    file_path = os.environ.get('EXCEL_FILE_PATH')
    
    db_params = {
        'dbname': db_name,
        'user': db_user,
        'password': db_password,
        'host': db_host,
        'port': db_port
    }
    
    # 3. Consulta SQL con el rango dinámico
    query = f"""
    SELECT 
    s.name AS "SECCIÓN",
    f.name AS "FAMILIA",
    sf.name AS "SUBFAMILIA",
    p.default_code AS "REFERENCIA",
    p.name AS "NOMBRE",
    p.default_code || '-' || p.name AS "ARTICULO",
    COALESCE(pm.name, '') AS "MARCA",
    CASE WHEN p.en_pruebas = true THEN 'Sí' ELSE 'No' END AS "PRODUCTO PROPIO",
    CASE WHEN p.obsoleto = true THEN 'Sí' ELSE 'No' END AS "ARTICULO OBSOLETO"
 
    FROM product_product p
    INNER JOIN product_category s ON s.id = p.seccion
    INNER JOIN product_category f ON f.id = p.familia
    INNER JOIN product_category sf ON sf.id = p.subfamilia
    LEFT JOIN product_marca pm ON pm.id = p.marca
    
    WHERE p.active
    ORDER BY p.default_code;
    """
    
    # 4. Ejecutar la consulta
    try:
        with psycopg2.connect(**db_params) as conn:
            with conn.cursor() as cur:
                cur.execute(query)
                resultados = cur.fetchall()
                headers = [desc[0] for desc in cur.description]
    except Exception as e:
        print(f"Error al conectar o ejecutar la consulta: {e}")
        sys.exit(1)
    
    if not resultados:
        print("No se obtuvieron resultados de la consulta.")
        return
    else:
        print(f"Se obtuvieron {len(resultados)} filas de la consulta.")
    
    # 5. Cargar el archivo base Portes.xlsx
    try:
        book = load_workbook(file_path)
        sheet = book.active
        existing_invoice_codes = {row[2] for row in sheet.iter_rows(min_row=2, values_only=True) if row[2] is not None}
    except FileNotFoundError:
        print(f"No se encontró el archivo base '{file_path}'. Se aborta para no perder el formato.")
        return

    # 6. Añadir nuevas filas sin duplicados
    for row in resultados:
        if row[2] not in existing_invoice_codes:
            sheet.append(row)
            new_row_index = sheet.max_row
            if new_row_index > 1:
                for col in range(1, sheet.max_column + 1):
                    source_cell = sheet.cell(row=new_row_index - 1, column=col)
                    target_cell = sheet.cell(row=new_row_index, column=col)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.alignment = copy.copy(source_cell.alignment)
    
    # 7. Actualizar referencia de la tabla "Productos"
    if "Productos" in sheet.tables:
        tabla = sheet.tables["Productos"]
        max_row = sheet.max_row
        max_col = sheet.max_column
        last_col_letter = get_column_letter(max_col)
        new_ref = f"A1:{last_col_letter}{max_row}"
        tabla.ref = new_ref
        print(f"Tabla 'Productos' actualizada a rango: {new_ref}")
    else:
        print("No se encontró la tabla 'Productos'. Se conservará el formato actual, pero no se actualizará la referencia de la tabla.")
    
    # 8. Guardar archivo
    book.save(file_path)
    print(f"Archivo guardado con la estructura de tabla en '{file_path}'.")
    
if __name__ == '__main__':
    main()